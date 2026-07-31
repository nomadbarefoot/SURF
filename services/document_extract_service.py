"""Safe document extraction for downloaded artifacts.

Extracts bounded text from common document formats. Rejects archives,
macros, embedded executables, and oversized documents.
"""
from __future__ import annotations

import csv
import json
import mimetypes
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from html.parser import HTMLParser
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import structlog

from config.settings import settings

logger = structlog.get_logger()

# Formats that are unconditionally rejected.
_REJECTED_EXTENSIONS: Set[str] = {
    "zip", "tar", "gz", "bz2", "xz", "7z", "rar", "jar", "apk",
    "exe", "dll", "bin", "sh", "bat", "cmd", "ps1", "vbs", "js",
    "doc", "xls", "ppt",  # legacy OLE/macro formats
    "iso", "img", "dmg",
}

_REJECTED_MIME_PREFIXES: Set[str] = {
    "application/zip",
    "application/x-zip",
    "application/x-tar",
    "application/x-gzip",
    "application/x-bzip2",
    "application/x-7z-compressed",
    "application/x-rar",
    "application/java-archive",
    "application/vnd.android.package-archive",
    "application/x-msdownload",
    "application/x-executable",
    "application/x-sh",
    "application/x-msdos-program",
    "application/x-iso9660-image",
}

_SUPPORTED_EXTENSIONS: Set[str] = {
    "txt", "md", "csv", "json", "html", "htm", "xml", "pdf", "docx", "xlsx",
}

# Content signatures for formats we refuse regardless of the declared name.
# The filename and Content-Type are attacker-influenced, so the bytes decide.
_MAGIC_SIGNATURES: List[tuple] = [
    (b"PK\x03\x04", "zip"),
    (b"PK\x05\x06", "zip"),
    (b"PK\x07\x08", "zip"),
    (b"MZ", "executable"),
    (b"\x7fELF", "executable"),
    (b"\xca\xfe\xba\xbe", "executable"),
    (b"\xcf\xfa\xed\xfe", "executable"),
    (b"\xce\xfa\xed\xfe", "executable"),
    (b"\x1f\x8b", "gzip"),
    (b"BZh", "bzip2"),
    (b"\xfd7zXZ", "xz"),
    (b"7z\xbc\xaf\x27\x1c", "7z"),
    (b"Rar!\x1a\x07", "rar"),
    (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1", "legacy-ole"),
    (b"#!", "script"),
]

# ZIP-container formats we do accept; their magic is indistinguishable from a
# plain archive, so the declared extension is what separates them.
_ZIP_BACKED_EXTENSIONS: Set[str] = {"docx", "xlsx"}

# Entity/DTD constructs that enable XXE and billion-laughs expansion.
_XML_UNSAFE = re.compile(rb"<!\s*(DOCTYPE|ENTITY)", re.IGNORECASE)


def _sniff_rejected_format(content: bytes, ext: str) -> Optional[str]:
    """Return a rejection label if *content* is a format we refuse.

    Guards against a benign extension (``report.txt``) wrapping an archive or
    executable payload.
    """
    head = content[:16]
    for signature, label in _MAGIC_SIGNATURES:
        if head.startswith(signature):
            if label == "zip" and ext in _ZIP_BACKED_EXTENSIONS:
                return None
            return label
    return None


class _TagStripper(HTMLParser):
    """Strip HTML tags and collapse whitespace."""

    def __init__(self):
        super().__init__()
        self.text = []

    def handle_data(self, data: str) -> None:
        self.text.append(data)

    def get_text(self) -> str:
        raw = "".join(self.text)
        return re.sub(r"\s+", " ", raw).strip()


def _strip_html(html: str) -> str:
    stripper = _TagStripper()
    try:
        stripper.feed(html)
        return stripper.get_text()
    except Exception:
        # Fallback to regex if the parser chokes.
        return re.sub(r"<[^>]+>", " ", html).strip()


@dataclass
class ExtractResult:
    """Outcome of document extraction."""

    success: bool
    content: str
    content_type: Optional[str]
    format: Optional[str]
    error: Optional[str] = None
    truncated: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "content": self.content,
            "content_type": self.content_type,
            "format": self.format,
            "error": self.error,
            "truncated": self.truncated,
        }


class DocumentExtractService:
    """Extract readable text from safe document artifacts."""

    def __init__(self, max_size_bytes: Optional[int] = None, max_text_length: Optional[int] = None):
        self.max_size_bytes = max_size_bytes or settings.max_download_size_bytes
        self.max_text_length = max_text_length or 8000

    def extract_from_path(
        self,
        path: str | Path,
        filename: Optional[str] = None,
        content_type: Optional[str] = None,
    ) -> ExtractResult:
        """Extract text from a file path after validating safety."""
        resolved = Path(path).resolve()
        if not resolved.is_file():
            return ExtractResult(False, "", content_type, None, "File not found")

        size = resolved.stat().st_size
        if size > self.max_size_bytes:
            return ExtractResult(
                False,
                "",
                content_type,
                None,
                f"Document exceeds size limit ({size} > {self.max_size_bytes})",
            )

        guessed_type, _ = mimetypes.guess_type(str(resolved))
        effective_type = content_type or guessed_type or "application/octet-stream"
        ext = (filename or resolved.name).rsplit(".", 1)[-1].lower()

        if ext in _REJECTED_EXTENSIONS:
            return ExtractResult(False, "", effective_type, ext, "Rejected file type")

        for prefix in _REJECTED_MIME_PREFIXES:
            if effective_type.lower().startswith(prefix):
                return ExtractResult(False, "", effective_type, ext, "Rejected MIME type")

        try:
            content = resolved.read_bytes()
            return self.extract_from_bytes(content, filename=filename or resolved.name, content_type=effective_type)
        except Exception as e:
            logger.error("Document extraction failed", path=str(resolved), error=str(e))
            return ExtractResult(False, "", effective_type, ext, str(e))

    def extract_from_bytes(
        self,
        content: bytes,
        filename: Optional[str] = None,
        content_type: Optional[str] = None,
    ) -> ExtractResult:
        """Extract text from raw bytes after validating safety."""
        if len(content) > self.max_size_bytes:
            return ExtractResult(
                False,
                "",
                content_type,
                None,
                f"Document exceeds size limit ({len(content)} > {self.max_size_bytes})",
            )

        ext = (filename or "download.bin").rsplit(".", 1)[-1].lower()
        guessed_type, _ = mimetypes.guess_type(filename or "download.bin")
        effective_type = content_type or guessed_type or "application/octet-stream"

        if ext in _REJECTED_EXTENSIONS:
            return ExtractResult(False, "", effective_type, ext, "Rejected file type")

        for prefix in _REJECTED_MIME_PREFIXES:
            if effective_type.lower().startswith(prefix):
                return ExtractResult(False, "", effective_type, ext, "Rejected MIME type")

        sniffed = _sniff_rejected_format(content, ext)
        if sniffed:
            return ExtractResult(
                False,
                "",
                effective_type,
                ext,
                f"Rejected file type (content is {sniffed}, declared .{ext})",
            )

        try:
            if ext in ("txt", "md") or effective_type.startswith("text/plain"):
                return self._extract_text(content, effective_type)
            if ext in ("html", "htm") or effective_type == "text/html":
                return self._extract_html(content, effective_type)
            if ext == "csv" or effective_type == "text/csv":
                return self._extract_csv(content, effective_type)
            if ext == "json" or effective_type == "application/json":
                return self._extract_json(content, effective_type)
            if ext == "xml" or effective_type.endswith("/xml") or effective_type.endswith("+xml"):
                return self._extract_xml(content, effective_type)
            if ext == "pdf" or effective_type == "application/pdf":
                return self._extract_pdf(content, effective_type)
            if ext == "docx":
                return self._extract_docx(content, effective_type)
            if ext == "xlsx":
                return self._extract_xlsx(content, effective_type)

            return ExtractResult(
                False,
                "",
                effective_type,
                ext,
                "Unsupported document format",
            )
        except Exception as e:
            logger.error("Document extraction failed", filename=filename, error=str(e))
            return ExtractResult(False, "", effective_type, ext, str(e))

    def _extract_text(self, content: bytes, content_type: str) -> ExtractResult:
        text = content.decode("utf-8", errors="replace")
        return self._ok(text, content_type, "text")

    def _extract_html(self, content: bytes, content_type: str) -> ExtractResult:
        html = content.decode("utf-8", errors="replace")
        text = _strip_html(html)
        return self._ok(text, content_type, "html")

    def _extract_csv(self, content: bytes, content_type: str) -> ExtractResult:
        text = content.decode("utf-8", errors="replace")
        rows = []
        try:
            reader = csv.reader(StringIO(text))
            for row in reader:
                rows.append(" | ".join(row))
        except Exception:
            rows = text.splitlines()
        return self._ok("\n".join(rows), content_type, "csv")

    def _extract_json(self, content: bytes, content_type: str) -> ExtractResult:
        data = json.loads(content.decode("utf-8", errors="replace"))
        text = json.dumps(data, ensure_ascii=False, indent=2)
        return self._ok(text, content_type, "json")

    def _extract_xml(self, content: bytes, content_type: str) -> ExtractResult:
        # ElementTree is vulnerable to entity-expansion DoS (billion laughs) and
        # will happily resolve internal entities, so refuse any DTD up front
        # rather than parsing it. Tag-stripping still recovers the text.
        if _XML_UNSAFE.search(content[:8192]):
            logger.warning("Rejected XML document declaring a DTD or entities")
            text = _strip_html(content.decode("utf-8", errors="replace"))
            return self._ok(text, content_type, "xml")

        text = content.decode("utf-8", errors="replace")
        try:
            root = ET.fromstring(text)
            text = " ".join(root.itertext())
            text = re.sub(r"\s+", " ", text).strip()
        except Exception:
            text = _strip_html(text)
        return self._ok(text, content_type, "xml")

    def _extract_pdf(self, content: bytes, content_type: str) -> ExtractResult:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            return ExtractResult(
                False,
                "",
                content_type,
                "pdf",
                f"PDF parser not installed: {exc}",
            )
        try:
            # PdfReader requires a binary stream; decoding to text corrupts the
            # xref table and breaks relative seeks.
            reader = PdfReader(BytesIO(content))
            pages = []
            for i, page in enumerate(reader.pages):
                try:
                    pages.append(page.extract_text() or "")
                except Exception as e:
                    logger.debug("PDF page extraction failed", page=i, error=str(e))
            text = "\n\n".join(pages)
            return self._ok(text, content_type, "pdf")
        except Exception as e:
            return ExtractResult(False, "", content_type, "pdf", str(e))

    def _extract_docx(self, content: bytes, content_type: str) -> ExtractResult:
        try:
            from docx import Document
        except ImportError as exc:
            return ExtractResult(
                False,
                "",
                content_type,
                "docx",
                f"DOCX parser not installed: {exc}",
            )
        try:
            doc = Document(BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text]
            text = "\n\n".join(paragraphs)
            return self._ok(text, content_type, "docx")
        except Exception as e:
            return ExtractResult(False, "", content_type, "docx", str(e))

    def _extract_xlsx(self, content: bytes, content_type: str) -> ExtractResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            return ExtractResult(
                False,
                "",
                content_type,
                "xlsx",
                f"XLSX parser not installed: {exc}",
            )
        try:
            wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
            rows = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
            text = "\n".join(rows)
            return self._ok(text, content_type, "xlsx")
        except Exception as e:
            return ExtractResult(False, "", content_type, "xlsx", str(e))

    def _ok(self, text: str, content_type: str, format: str) -> ExtractResult:
        truncated = len(text) > self.max_text_length
        return ExtractResult(
            success=True,
            content=text[: self.max_text_length],
            content_type=content_type,
            format=format,
            truncated=truncated,
        )


_document_extract_service: Optional[DocumentExtractService] = None


def get_document_extract_service() -> DocumentExtractService:
    """Lazy singleton for DocumentExtractService."""
    global _document_extract_service
    if _document_extract_service is None:
        _document_extract_service = DocumentExtractService()
    return _document_extract_service


def reset_document_extract_service() -> None:
    """Reset singleton, mainly for tests."""
    global _document_extract_service
    _document_extract_service = None
